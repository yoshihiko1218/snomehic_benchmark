#!/usr/bin/env python
"""Build an Excel summary of all datasets used in the scNOME-HiC benchmark.

Source of truth: the per-folder markdowns (FILES.md / RESULTS.md), the top-level
README.md, PROJECT_CONTEXT.md, and analysis_pipeline.md. This script hard-codes the
distilled facts from those docs so the workbook is reproducible.

Output: summary/benchmark_data_summary.xlsx
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), "benchmark_data_summary.xlsx")

# ---------------------------------------------------------------------------
# Sheet 1 — Datasets (one row per benchmark dataset folder)
# ---------------------------------------------------------------------------
DATASET_COLS = [
    "Folder", "Method / technology", "Publication (source)", "Genome",
    "Cell type / tissue", "Modalities",
    "Cells listed", "Cells used in benchmark",
    "SRA / GEO accessions", "Pipeline / aligner", "Role & notes",
]

DATASETS = [
    ["scnomehic/", "scNOMe-HiC (this method)", "This study", "hg38",
     "GM12878 (lymphoblastoid cell line)", "CpG meth (HCG) + GpC accessibility (GCH) + Hi-C",
     "188", "187 (QC-pass)",
     "In-house (scNH_GM_4plex_*); bhmem output external at b1198 gm_sc_new",
     "TrimGalore -> bhmem (bisulfitehic BWA-MEM) -> markdup/calmd; YAP bowtie2 arm as local comparison",
     "The published method being benchmarked. Three modalities. bhmem is the method aligner (external); local YAP/bowtie2 arm feeds mapping/Hi-C QC."],

    ["nagano/", "Nagano scHi-C", "Nagano et al. 2013", "mm10",
     "Mouse ESC (single cells)", "Hi-C only (no methylation)",
     "15", "15",
     "SRR921526-SRR921540", "fastp -> bwa-mem2 -SP5M -> samtools markdup -> pairtools",
     "Early pure scHi-C reference. High trans_ratio (~0.17) is a real quality signature of this early assay."],

    ["droplethic/", "Droplet Hi-C", "Chang et al. 2025", "hg38",
     "Cell line (barcoded droplets)", "Hi-C only",
     "~688k barcodes", "3,668 valid barcodes",
     "SRR27586278, SRR27586279", "Rupture: TrimGalore -> bowtie1 barcode -> bwa mem -SP5M -> CB-tag -> pairtools",
     "High-throughput droplet Hi-C. CAVEAT: pairtools dedup not run -> values are currently pre-dedup (kept as-is by decision 2026-06-30)."],

    ["scnome/", "scNOMe-seq", "Pott 2017", "hg38",
     "GM12878 + K562 (cell lines)", "CpG meth (HCG) + GpC accessibility (GCH); no Hi-C",
     "34 runs (after control removal)", "23 cells (12 GM12878 + 11 merged K562)",
     "SRR3729642-SRR3729653 (GM12878); SRR3729661-SRR3729682 (K562); SRR3729654-3729660 = excluded spike-in controls",
     "TrimGalore -> Bismark SE per-mate (--non_directional) -> markdup -> coverage2cytosine --nome-seq",
     "NOMe method (CpG + GCH accessibility). K562 has 2 runs/cell merged into 11 cells; controls excluded."],

    ["snmCseq2/", "snmC-seq2", "Luo 2018", "hg38 + mm10",
     "Mixed species single cells", "CpG methylation only (not NOMe, no GCH); no Hi-C",
     "249 (153 hg38 + 96 mm10)", "mm10 subset = 96",
     "SRR6911624 ... (250-line acc_list)", "cutadapt -> Bismark SE per-mate -> dedup -> methyl-extractor; parallel YAP mc arm",
     "CpG-only methylome. mm10 subset used for the benchmark figures."],

    ["snmCseq3/", "snm3C-seq", "Liu 2023", "mm10",
     "Mouse single cells", "CpG methylation + Hi-C (m3C chimeric split-read contacts)",
     "1,379 in SRA pool (100 aligned)", "98",
     "SRR21549289 ... (1379-line acc_list)", "cutadapt -> YAP bismark+bowtie2 m3C (alignment/) AND bhmem (04.bhmem_bam/)",
     "Combined methylome + Hi-C. Final figures: align+contacts from YAP, conversion+HCG loci from bhmem. Contacts are chimeric multi-way (method-inherent difference)."],

    ["smallwood/", "Smallwood scBS-seq (scWGBS)", "Smallwood 2014", "mm10",
     "Mouse ESC (single cells)", "CpG methylation only (WGBS); no Hi-C, no GCH",
     "51", "51",
     "SRR1248444 ... (51-line acc_list)", "TrimGalore -> Bismark (hg38 contamination depletion -> mm10 SE) -> markdup -> cov + BisSNP",
     "Whole-genome bisulfite single-cell reference (no accessibility, no Hi-C)."],

    ["snmCAT/", "snmC2T-seq / snmCAT-seq (NOMe)", "Luo 2022", "hg38",
     "Human frontal cortex, donor UMB5580 (batch 190321)", "CpG meth (HCG) + GpC accessibility (GCH) + RNA (transcriptome)",
     "100", "99",
     "GSE140493 (SRR* per cells_brain.tsv)", "YAP mct --nome (bismark methylome + STAR transcriptome + allcools)",
     "Multi-omic NOMe brain data. NB: autosomal non-CpG is elevated by real brain neuronal mCH, not conversion failure. Median HCG 81.2% / GCH 15.1%."],

    ["methylhic/", "Methyl-HiC", "Li et al. 2019", "mm10",
     "Single cells", "CpG methylation + Hi-C",
     "59", "Excluded from current figures",
     "SRR7770822 ... (59-line acc_list)", "YAP (m3C)",
     "Excluded from the current benchmark figures per user decision."],

    ["methylhic_new/", "Methyl-HiC (newer batch)", "Li et al. 2019 (newer batch)", "mm10",
     "Single cells", "CpG methylation + Hi-C",
     "96", "Excluded from current figures",
     "In-house sample IDs (sc_*_CKDL*, 96-line acc_list)", "Snakemake per-Group alignment",
     "Newer Methyl-HiC batch; also excluded from current benchmark figures per user decision."],
]

# ---------------------------------------------------------------------------
# Sheet 2 — QC metrics measured across datasets
# ---------------------------------------------------------------------------
METRIC_COLS = ["#", "Metric", "Definition (held constant across methods)", "Applies to"]
METRICS = [
    ["1", "Bisulfite conversion", "Non-CpG methylation % in ACT trinucleotide context (H=A/C/T), for chrM and an autosome (chr21 human / chr19 mouse). Lower = better conversion.", "All bisulfite methods"],
    ["2", "Uniquely-mapped count", "Fragments (R1+R2 collapsed to one read-name) uniquely mapped; reported both pre- and post-dedup.", "All"],
    ["3", "MapQ30 mapping rate", "MapQ>=30 fragments / mapped fragments; reported pre- and post-dedup.", "All"],
    ["4", "Per-cell cis contacts", "MapQ>=30, deduplicated cis Hi-C contacts (cis_n = all cis).", "Hi-C methods"],
    ["5", "Trans / (cis+trans) ratio", "trans / (cis + trans) at MapQ>=30, deduplicated.", "Hi-C methods"],
    ["6", "Per-cell cis >1kb contacts", "Cis contacts with |pos1-pos2| > 1 kb (cis-long threshold = 1 kb).", "Hi-C methods"],
    ["-", "HCG detected loci", ">=1x-covered CpG cytosine sites, destranded, GCG removed by reference lookup.", "Methylation methods"],
    ["-", "GCH detected loci", ">=1x-covered GpC cytosine sites, destranded, GCG removed. NOMe accessibility.", "NOMe methods (scnomehic, scnome, snmCAT)"],
]

# ---------------------------------------------------------------------------
# Sheet 3 — Known method-inherent caveats
# ---------------------------------------------------------------------------
CAVEAT_COLS = ["#", "Caveat"]
CAVEATS = [
    ["1", "Hi-C contact model differs: snm3C-seq = chimeric split-read (multi-way); nagano / scnomehic / droplethic = paired-mate. Thresholds (1kb / MapQ30 / dedup) are matched."],
    ["2", "snmCAT autosomal non-CpG conversion is confounded by real brain neuronal mCH (not a conversion failure)."],
    ["3", "Conversion autosome proxy is chr21 (human) vs chr19 (mouse)."],
    ["4", "droplethic is not yet PCR-deduplicated -> its metrics are pre-dedup (before == after)."],
    ["5", "methylhic and methylhic_new are excluded from current benchmark figures per user decision."],
]

# ---------------------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E78")
THIS_METHOD_FILL = PatternFill("solid", fgColor="FCE4D6")
EXCLUDED_FILL = PatternFill("solid", fgColor="EDEDED")
ALT_FILL = PatternFill("solid", fgColor="F2F7FB")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_sheet(ws, title, columns, rows, widths, wrap_cols, row_fill=None):
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    hdr = 3
    for j, col in enumerate(columns, start=1):
        c = ws.cell(row=hdr, column=j, value=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    for i, row in enumerate(rows):
        r = hdr + 1 + i
        fill = row_fill(row, i) if row_fill else (ALT_FILL if i % 2 else None)
        for j, val in enumerate(row, start=1):
            c = ws.cell(row=r, column=j, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=(j - 1) in wrap_cols)
            c.border = BORDER
            if fill:
                c.fill = fill
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    ws.row_dimensions[hdr].height = 30


def dataset_fill(row, i):
    if "this method" in row[1].lower():
        return THIS_METHOD_FILL
    if "excluded" in str(row[7]).lower():
        return EXCLUDED_FILL
    return ALT_FILL if i % 2 else None


wb = Workbook()

ws1 = wb.active
ws1.title = "Datasets"
write_sheet(
    ws1, "scNOME-HiC Benchmark - Datasets used", DATASET_COLS, DATASETS,
    widths=[15, 26, 22, 13, 26, 34, 16, 20, 40, 46, 60],
    wrap_cols={1, 2, 4, 5, 8, 9, 10},
    row_fill=dataset_fill,
)
# legend note under the table
note_r = 3 + 1 + len(DATASETS) + 1
ws1.cell(row=note_r, column=1,
         value="Legend: orange = the published method (scNOME-HiC); grey = present but excluded from current figures. "
               "Cell counts and provenance are distilled from each folder's FILES.md / RESULTS.md and analysis_pipeline.md."
         ).font = Font(italic=True, size=9, color="808080")

ws2 = wb.create_sheet("QC metrics")
write_sheet(
    ws2, "QC metrics measured across the benchmark", METRIC_COLS, METRICS,
    widths=[5, 26, 78, 40], wrap_cols={2, 3},
)

ws3 = wb.create_sheet("Caveats")
write_sheet(
    ws3, "Known method-inherent caveats", CAVEAT_COLS, CAVEATS,
    widths=[5, 120], wrap_cols={1},
)

wb.save(OUT)
print("Wrote", OUT)
print("Sheets:", wb.sheetnames)
print("Datasets:", len(DATASETS))
